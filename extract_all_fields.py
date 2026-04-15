import openpyxl

def extract_all_fields(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    field_positions = {
        'C': 'PRODUCT NAME',
        'D': 'MATERIAL',
        'E': 'COLOR',
        'F': 'PRODUCT SIZE',
        'G': 'PACKAGING SIZE',
        'H': 'FRENCH DESC',
        'I': 'THEME',
        'J': 'BARCODE',
        'K': 'HS CODE',
        'L': 'SUPPLIER CODE',
    }

    # Add header row names to row 1
    for col, field_name in field_positions.items():
        ws[f'{col}1'].value = field_name

    for row in range(2, ws.max_row + 1):
        b_cell = f'B{row}'
        content = ws[b_cell].value

        # Skip non-string cells (dates, numbers, blanks)
        if not isinstance(content, str):
            continue

        # Split content into lines
        lines = content.split('\n')
        
        if len(lines) == 0:
            continue
        
        # Extract first line = product name (copy to column C)
        first_line = lines[0].replace('_x000D_', '').strip()
        ws[f'C{row}'].value = first_line
        
        # Remove first line from column B (remaining lines move up)
        remaining_lines = lines[1:]  # All lines except the first
        remaining_content = '\n'.join(remaining_lines)
        
        # Update column B with remaining content (second line becomes first)
        ws[b_cell].value = remaining_content
        
        # Process remaining content for other fields
        for line in remaining_lines:
            clean_line = line.replace('_x000D_', '').strip()
            for col, field_name in field_positions.items():
                if field_name != 'PRODUCT NAME' and field_name.upper() in clean_line.upper() and ':' in clean_line:
                    value = clean_line.split(':', 1)[1].strip()
                    ws[f'{col}{row}'].value = value
                    break

    wb.save(filename)

extract_all_fields("iba_prodml_export_memo_clean1.xlsx")


