"""
Management command to import products from Excel template
"""
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from openpyxl import load_workbook
from qa_app.models import Product, ProductStage, ComplianceDocument
from datetime import datetime
import os


class Command(BaseCommand):
    help = 'Import products from Excel template (MTL sheet format)'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Path to Excel file')
        parser.add_argument(
            '--sheet',
            type=str,
            default='MTL (2)',
            help='Sheet name to import from (default: MTL (2))'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Dry run - do not save to database'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        sheet_name = options['sheet']
        dry_run = options['dry_run']

        if not os.path.exists(excel_file):
            self.stdout.write(self.style.ERROR(f'File not found: {excel_file}'))
            return

        self.stdout.write(f'Loading Excel file: {excel_file}')
        wb = load_workbook(excel_file, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            self.stdout.write(self.style.ERROR(f'Sheet "{sheet_name}" not found'))
            self.stdout.write(f'Available sheets: {", ".join(wb.sheetnames)}')
            return

        ws = wb[sheet_name]
        
        # Find header row (assuming first row contains headers)
        headers = {}
        header_row = 1
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx

        self.stdout.write(f'Found {len(headers)} columns')
        
        # Get default user for created_by
        default_user = User.objects.first()
        if not default_user:
            self.stdout.write(self.style.ERROR('No users found. Please create a user first.'))
            return

        imported_count = 0
        skipped_count = 0

        # Process rows (skip header row)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # Check if row has data (BMUK Item No. is required)
            bmuk_item_no = self.get_cell_value(row, headers.get('BMUK Item No.'))
            if not bmuk_item_no:
                continue

            try:
                # Create or update product
                product_data = {
                    'bmuk_item_no': str(bmuk_item_no).strip(),
                    'mtl_ref_no': str(self.get_cell_value(row, headers.get('MTL Ref NO.')) or '').strip(),
                    'prism_code': str(self.get_cell_value(row, headers.get('PRISM')) or '').strip(),
                    'sub_category': str(self.get_cell_value(row, headers.get('Sub Cat')) or '').strip(),
                    'description': str(self.get_cell_value(row, headers.get('Description')) or '').strip(),
                    'product_specification': str(self.get_cell_value(row, headers.get('Product Specification')) or '').strip(),
                    'care_information': str(self.get_cell_value(row, headers.get('Care Information')) or '').strip(),
                    'product_category': str(self.get_cell_value(row, headers.get('Age Grade')) or '').strip(),
                    'material_type': self.map_material_type(self.get_cell_value(row, headers.get('Material'))),
                    'new_repeat_status': self.map_new_repeat_status(self.get_cell_value(row, headers.get('New / Repeat'))),
                    'supplier_code': str(self.get_cell_value(row, headers.get('Supplier Code')) or '').strip(),
                    'supplier_name': str(self.get_cell_value(row, headers.get('Supplier name')) or '').strip(),
                    'factory_item_no': str(self.get_cell_value(row, headers.get('supplier Item No.')) or '').strip(),
                    'bm_fr_item_no': str(self.get_cell_value(row, headers.get('BM FR item No.')) or '').strip(),
                    'fob_port': str(self.get_cell_value(row, headers.get('FOB Port')) or '').strip(),
                    'merchandiser_name': str(self.get_cell_value(row, headers.get('Merchandiser')) or '').strip(),
                    'test_requirements': str(self.get_cell_value(row, headers.get('Test Requirements')) or '').strip(),
                    'merchant_enquiry_date': self.parse_date(self.get_cell_value(row, headers.get('Merchant Enquiry date'))),
                    'shipdate_crd': self.parse_date(self.get_cell_value(row, headers.get('Shipdate CRD'))),
                    'created_by': default_user,
                    'status': 'in_progress',
                }

                if not product_data['mtl_ref_no']:
                    product_data['mtl_ref_no'] = product_data['bmuk_item_no']

                if dry_run:
                    self.stdout.write(f'[DRY RUN] Would create/update: {product_data["bmuk_item_no"]}')
                else:
                    product, created = Product.objects.update_or_create(
                        bmuk_item_no=product_data['bmuk_item_no'],
                        defaults=product_data
                    )
                    
                    # Create stages if they don't exist
                    for stage_code, stage_name in ProductStage.STAGE_CHOICES:
                        ProductStage.objects.get_or_create(
                            product=product,
                            stage_type=stage_code,
                            defaults={'status': 'not_started'}
                        )
                    
                    # Update stage completion dates if provided
                    stage_dates = {
                        'R': self.parse_date(self.get_cell_value(row, headers.get('Test Plan 1st issued date'))),
                        'A': self.parse_date(self.get_cell_value(row, headers.get('Artwork Reviewed date'))),
                        'F': self.parse_date(self.get_cell_value(row, headers.get('Factory Sample check date'))),
                        'M': self.parse_date(self.get_cell_value(row, headers.get('Mockup / Red sample check date'))),
                        'G': self.parse_date(self.get_cell_value(row, headers.get('Shipment / Gold Seal sample check date'))),
                    }
                    
                    for stage_code, completion_date in stage_dates.items():
                        if completion_date:
                            stage, _ = ProductStage.objects.get_or_create(
                                product=product,
                                stage_type=stage_code
                            )
                            stage.completion_date = completion_date
                            stage.status = 'completed'
                            if not stage.start_date:
                                stage.start_date = completion_date
                            stage.save()
                            
                            # Update artwork status if available
                            if stage_code == 'A':
                                artwork_status = self.get_cell_value(row, headers.get('Artwork Status'))
                                if artwork_status:
                                    stage.notes = str(artwork_status).strip()
                                    stage.save()

                    imported_count += 1
                    action = 'Created' if created else 'Updated'
                    self.stdout.write(self.style.SUCCESS(f'{action}: {product.bmuk_item_no}'))

            except Exception as e:
                skipped_count += 1
                self.stdout.write(self.style.ERROR(f'Error processing row {row_idx}: {str(e)}'))
                continue

        self.stdout.write(self.style.SUCCESS(
            f'\nImport complete. Imported: {imported_count}, Skipped: {skipped_count}'
        ))

    def get_cell_value(self, row, col_idx):
        """Get cell value by column index (1-based)"""
        if not col_idx or col_idx > len(row):
            return None
        return row[col_idx - 1].value

    def parse_date(self, value):
        """Parse date from Excel cell"""
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, str):
            try:
                return datetime.strptime(value, '%Y-%m-%d').date()
            except:
                pass
        return None

    def map_material_type(self, value):
        """Map material type from Excel to model choice"""
        if not value:
            return 'other'
        value = str(value).lower().strip()
        mapping = {
            'plastic': 'plastic',
            'fabric': 'fabric',
            'cosmetic': 'cosmetic',
            'wood': 'wood',
            'slime': 'slime',
            'paper': 'paper',
            'paint': 'paint',
            'crayon': 'crayon',
            'dough': 'dough',
        }
        return mapping.get(value, 'other')

    def map_new_repeat_status(self, value):
        """Map new/repeat status from Excel to model choice"""
        if not value:
            return 'new_item'
        value = str(value).lower().strip()
        if 'repeat' in value:
            return 'repeat'
        if 'new' in value:
            return 'new_item'
        return 'new_item'

